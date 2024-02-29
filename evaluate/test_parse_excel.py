import os
import openpyxl
import re
import __init__
import tools.utils
from tools.timer import Timer

from evaluate.db_operate_specification import DBOperateSpecification


from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

from tools.logger import getLogger

logger = getLogger()

# TEST_ASSET_FILE_NAME = 'AI眼镜规范查询测试汇总.xlsx'
# TEST_ASSET_FILE_NAME = 'AI眼镜规范查询测试汇总_2.0.xlsx'
TEST_ASSET_FILE_NAME = '规范检索_问答_1.0.xlsx'


class ExcelOperate():
    START_DATA_ROW = 3  # 开始的数据行

    # 最大线程数，如果要使用单线程，设置为1
    MAX_THREADS = 10

    def __init__(self, file_path):  # sourcery skip: raise-specific-error
        self.file_path = file_path
        super().__init__()

    # 读取 Excel 表格
    def read_excel(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        data = []
        row_counter = 1  # 计数器，用于跳过前两行
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_counter >= self.START_DATA_ROW and row[0]:
                # 这里的 row_num 是行号
                data.append((row_num, {
                    '隐患': row[0],
                    '答案': row[1],
                    '相似答案': row[2],
                    # Add more columns as needed
                }))
            row_counter += 1

        return data

    def get_col_num_b_column_name(self, col_name):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        # 获取列名所在的列号
        column_number = None
        for col_num, header in enumerate(sheet[1], start=1):
            if header.value == col_name:
                column_number = col_num
                break

        if column_number is None:
            logger.warning(f"找不到列名为 '{col_name}' 的列。")

        return column_number

    def update_cell_by_column_name(self, row_num, col_name, new_value):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        # 获取列名所在的列号
        column_number = None
        for col_num, header in enumerate(sheet[1], start=1):
            if header.value == col_name:
                column_number = col_num
                break

        if column_number is not None:
            # 获取要更新的单元格对象
            cell = sheet.cell(row=row_num, column=column_number)

            # 更新单元格的值
            cell.value = new_value

            # 保存到文件
            wb.save(self.file_path)
            logger.info("单元格的值已更新。")
        else:
            logger.warning(f"找不到列名为 '{col_name}' 的列。")

    def update_cell(self, row_num, col_num, new_value):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        # 获取要更新的单元格对象
        cell = sheet.cell(row=row_num, column=col_num)

        # 更新单元格的值
        cell.value = new_value

        # 保存到文件
        wb.save(self.file_path)

    def batch_update_cell(self, cell_updates):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        for cell_update in cell_updates:
            row_num = cell_update[0]
            col_num = cell_update[1]
            new_value = cell_update[2]

            # 获取要更新的单元格对象
            cell = sheet.cell(row=row_num, column=col_num)

            # 更新单元格的值
            cell.value = new_value

        # 保存到文件
        wb.save(self.file_path)

    # 使用正则表达式匹配 x.x.x 格式的前缀
    def extract_prefix(self, input_string):

        match = re.match(r'^\d+\.\d+\.\d+', input_string)

        if match:
            return match.group()
        else:
            return None

    # row_number, row_question两个参数，仅为了错误时输出日志用
    def calc_value(self, row_number, row_question, query_result, row_answer):
        try:
            # 数据库中找不到答案，设置为50，输出警告，一般情况下，50的应该是需要删除的，属于非行业问题
            if query_result is None:
                logger.warning(f"query_result is None in calc_value,row_number:{row_number},row_question：{row_question}")
                return 50

            code_expect = self.extract_prefix(row_answer)

            for index, result in enumerate(query_result, start=1):
                result_content = result[3]  # 规范原文
                code_result = self.extract_prefix(result_content)
                if code_result == code_expect:
                    return index
            return 20
        except Exception as e:
            logger.error(f"exception in calc_value,row_number:{row_number},row_question：{row_question},exception:{e}")
            return 30  # 异常情况用30，可以在结果中查看问题所在

    def process_row(self, row_number, row_data, column_name, compare_vector_field, compare_vector_operator, finetune_model_path):
        dbOperateSpecification = DBOperateSpecification()  # 创建数据库操作实例

        # 处理单行的逻辑...
        # 注意：在多线程环境中，所有涉及到修改共享资源的操作都需要加锁
        # 分别读取隐患，答案
        row_question = row_data['隐患']
        row_answer = row_data['答案']
        row_answer2 = row_data['相似答案']

        if (row_answer is not None and len(row_answer.strip()) > 0):
            query_result = dbOperateSpecification.kb_query(
                row_question, '', compare_vector_field, compare_vector_operator, 0.6, 15, finetune_model_path)
            cell_value = self.calc_value(row_number, row_question, query_result, row_answer)
        else:
            cell_value = None

        # 检测相似答案，取计算的最小值
        if (row_answer2 is not None and len(row_answer2.strip()) > 0):
            query_result = dbOperateSpecification.kb_query(
                row_question, '', compare_vector_field, compare_vector_operator, 0.6, 15, finetune_model_path)
            cell_value2 = self.calc_value(row_number, row_question, query_result, row_answer2)
            if cell_value2 is not None and cell_value is not None and cell_value2 < cell_value:
                log_content = f'__test__({column_name}),相似答案优于答案，选相似答案,cell_value={cell_value},cell_value2={cell_value2},row_number:{row_number}'
                print(log_content)
                cell_value = cell_value2
            elif cell_value2 is not None and cell_value is None:
                cell_value = cell_value2
                log_content = f'__test__({column_name}),无答案，选相似答案,cell_value={cell_value},row_number:{row_number}'
                print(log_content)

        # 返回单元格值和行号，以便稍后更新Excel
        return (row_number, cell_value)

    def __test__(self, compare_vector_field, compare_vector_operator, finetune_model_path):
        timer = Timer()

        # 因为数据库支持正式和训练时使用不同的字段名称，excel不想做特殊处理，在这里进行替换
        compare_vector_field_name = compare_vector_field.replace("_train", "")
        column_name = f"{compare_vector_field_name}({compare_vector_operator})"   # 列名规则规定
        timer.text = f"__test__({column_name})," + "Elapsed time: {:0.4f} seconds"
        timer.start()

        # 读取 Excel 表格
        read_data = self.read_excel()
        column_number = self.get_col_num_b_column_name(column_name)

        update_lock = threading.Lock()  # 创建锁以同步Excel文件的更新

        cell_values = []  # List to store all cell_value instances
        cell_updates = []  # 先保存结果，后面一次性写入excel中
        handle_count = 0
        total_count = len(read_data)
        last_progress = 0

        # 创建一个线程池
        with ThreadPoolExecutor(max_workers=self.MAX_THREADS) as executor:
            # 提交所有行数据到线程池中处理
            futures = [executor.submit(self.process_row, row_number, row_data, column_name, compare_vector_field, compare_vector_operator,
                                       finetune_model_path) for row_number, row_data in read_data]

            for future in as_completed(futures):
                try:
                    # row_data = futures[future]
                    row_number, cell_value = future.result()
                except Exception as e:
                    print(f"Row processing generated an exception: {e}")
                    cell_value = None

                # 在这里可以更新Excel，但它必须是线程安全的
                with update_lock:  # 使用锁来保证线程安全性
                    if cell_value is not None:
                        cell_values.append(cell_value)
                        cell_updates.append((row_number, column_number, cell_value))

                    # 控制进度日志，百分比有变化且是5的倍数输出
                    handle_count += 1
                    current_progress = handle_count * 100 // total_count
                    if last_progress != current_progress and (current_progress % 5 == 0):
                        log_content = f'__test__({column_name})处理进度：{handle_count}/{total_count},百分比：{current_progress}%'
                        print(f"{Timer.current_time_str()}:{log_content}")
                        last_progress = current_progress

        # 在循环外部，一次性写入Excel
        self.batch_update_cell(cell_updates)

        # 计算结果
        results = self.calculate_results(cell_values)

        timer.stop()

        return results

    def calculate_results(self, cell_values):
        # 异常处理，赋默认值，避免报错
        if not cell_values:
            results = {
                'average': 20.0,
                'top1': 0.0,
                'top3': 0.0,
                'top5': 0.0,
                'top10': 0.0,
                'top15': 0.0
            }
            return results

        # Calculate average
        average_value = sum(cell_values) / len(cell_values)

        # Calculate proportions
        def calculate_proportion(n):
            return sum(value <= n for value in cell_values) / len(cell_values)

        # Getting top N metrics
        results = {
            'average': average_value,
            'top1': calculate_proportion(1),
            'top3': calculate_proportion(3),
            'top5': calculate_proportion(5),
            'top10': calculate_proportion(10),
            'top15': calculate_proportion(15)
        }

        return results

    def print_results(self, train_params, all_results):
        def log_results(msg):
            # print(msg)
            logger.warning(msg)

        headers = ["compare_vector_field", "Average", "Top 1", "Top 3", "Top 5", "Top 10", "Top 15"]
        data_rows = []

        for key, metrics in all_results.items():
            field_name, operator = key.split('(')
            operator = operator.strip(')')
            row = [
                f"{field_name}({operator})",
                f"{metrics['average']:.2f}",
                f"{metrics['top1']:.2%}",
                f"{metrics['top3']:.2%}",
                f"{metrics['top5']:.2%}",
                f"{metrics['top10']:.2%}",
                f"{metrics['top15']:.2%}"
            ]
            data_rows.append(row)

        # Calculate column widths
        column_widths = [max(len(str(cell)) for cell in col_cells) for col_cells in zip(*([headers] + data_rows))]

        # Create and print header row
        header_row = "| " + " | ".join(header.ljust(width) for header, width in zip(headers, column_widths)) + " |"
        line = "+" + "+".join("-" * (width + 2) for width in column_widths) + "+"

        results = []
        results.append(line)
        results.append(header_row)
        results.append(line)

        # Print data rows
        for row in data_rows:
            results.append("| " + " | ".join(cell.ljust(width) for cell, width in zip(row, column_widths)) + " |")
            results.append(line)

        # 将结果列表中的每个元素合并为单个字符串
        result_str = "\n".join(results)
        log_results(
            f"train_params:({train_params}),test_asset_file_name:({TEST_ASSET_FILE_NAME}),evaluate results as follows:\r\n" + result_str)


def __make_vector_content(train_params, compare_vector_fields, finetune_model_path=""):
    # 获取当前文件所在的目录
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # 如果指定了微调模型的路径，复制一个excel文件进行再进行操作（如果文件已经存在，则复用，主要是为解决单次微调后结果保存在一次的目的）
    if finetune_model_path:
        import shutil
        from pathlib import Path

        path = Path(finetune_model_path)
        last_folder_name = path.name
        excel_file_path = os.path.join(current_dir, "test-asset", f"{last_folder_name}_{TEST_ASSET_FILE_NAME}")
        if not os.path.exists(excel_file_path):
            source_file = os.path.join(current_dir, "test-asset", TEST_ASSET_FILE_NAME)
            shutil.copy(source_file, excel_file_path)
    else:
        excel_file_path = os.path.join(current_dir, "test-asset", TEST_ASSET_FILE_NAME)

    excelOperate = ExcelOperate(excel_file_path)

    all_results = {}

    # 负内积 (Negative Dot Product) <#>   余弦 (Cosine): cosine_similarity <=>  欧几里得距离 (Euclidean Distance) <->
    # 为缩短验证时间，先取一个，其他的关闭，建议使用 余弦  或 欧几里得距离
    negative_dot_product = False
    cosine_similarity = True
    euclidean_distance = True

    if not negative_dot_product and not cosine_similarity and not euclidean_distance:
        raise Exception("negative_dot_product,cosine_similarity,euclidean_distance至少设置一个为True")

    for compare_vector_field in compare_vector_fields:
        if negative_dot_product:
            results = excelOperate.__test__(compare_vector_field, '<#>', finetune_model_path)
            all_results[f"{compare_vector_field}(<#>)"] = results

        if cosine_similarity:
            results_2 = excelOperate.__test__(compare_vector_field, '<=>', finetune_model_path)
            all_results[f"{compare_vector_field}(<=>)"] = results_2

        if euclidean_distance:
            results_3 = excelOperate.__test__(compare_vector_field, '<->', finetune_model_path)
            all_results[f"{compare_vector_field}(<->)"] = results_3

    excelOperate.print_results(train_params, all_results)


def make_m3e_vector_content(train_params, finetune_model_path=""):
    __make_vector_content(train_params, ['m3e_vector_content'], finetune_model_path)


def make_pms_m3e_vector_content(train_params, finetune_model_path):
    # 如果数据集已经做过一次基础模型的评估，就不要加入了，当前数据集多了很多，评估会很慢 2023.12.19 lihongquan
    __make_vector_content(train_params, ['m3e_vector_content', DBOperateSpecification.DB_NAME_PMS_M3E_VECTOR], finetune_model_path)
    # __make_vector_content(train_params, [DBOperateSpecification.DB_NAME_PMS_M3E_VECTOR], finetune_model_path)


def make_bge_large_vector_content(train_params, finetune_model_path=""):
    __make_vector_content(train_params, ['bge_large_vector_content'], finetune_model_path)


def make_pms_bge_large_vector_content(train_params, finetune_model_path=""):
    # 如果数据集已经做过一次基础模型的评估，就不要加入了，当前数据集多了很多，评估会很慢 2023.12.19 lihongquan
    __make_vector_content(train_params, ['bge_large_vector_content',
                          DBOperateSpecification.DB_NAME_PMS_BGE_LARGE_VECTOR], finetune_model_path)
    # __make_vector_content(train_params, [DBOperateSpecification.DB_NAME_PMS_BGE_LARGE_VECTOR], finetune_model_path)


if __name__ == '__main__':  # sourcery skip: raise-specific-error
    train_params = "ManualTest"
    finetune_model_path = '/alidata/models/BAAI/bge-large-zh-v1.5-finetune-v2.0-train/datasets_datasets_v2.1_epochs_5_batch8_lr1e-05_test_0.2'

    # make_m3e_vector_content(train_params, finetune_model_path)
    make_bge_large_vector_content(train_params, finetune_model_path)

    # make_pms_m3e_vector_content(train_params, finetune_model_path)
    # make_pms_bge_large_vector_content(train_params, finetune_model_path)

    pass
